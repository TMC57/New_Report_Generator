"""
Parser pour le fichier Excel "Listing clients Orsye-Wash.xlsx"
Extrait les informations nécessaires pour enrichir les rapports
"""

import openpyxl
from typing import Dict, List, Optional
from pathlib import Path


def parse_listing_clients_excel(file_path: str) -> Dict[int, dict]:
    """
    Parse le fichier Excel "Listing clients Orsye-Wash.xlsx"
    
    Colonnes attendues:
    - Date installation
    - N° client
    - Client (automatique)
    - Groupe (automatique)
    - Date installation (duplicate?)
    - N° de zone de lavage
    - Adresse (automatique)
    - N° de routeur
    - Date dernière intervention
    - Produit lavant
    
    Returns:
        Dict[facility_id, dict]: Dictionnaire avec facility_id comme clé
    """
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Lire les en-têtes
    headers = [cell.value for cell in ws[1]]
    
    # Mapping des données par facility_id (N° client)
    facilities_data = {}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):  # Skip empty rows
            continue
            
        # Créer un dictionnaire pour cette ligne
        row_data = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_data[header] = row[i]
        
        # Extraire le facility_id (N° client)
        facility_id = row_data.get('N° client')
        
        if facility_id:
            # Convertir en int si possible
            try:
                facility_id = int(facility_id)
            except (ValueError, TypeError):
                continue
            
            # Stocker les informations utiles
            facilities_data[facility_id] = {
                'client_name': row_data.get('Client (automatique)', ''),
                'group': row_data.get('Groupe (automatique)', ''),
                'installation_date': row_data.get('Date installation'),
                'zone_number': row_data.get('N° de zone de lavage'),
                'address': row_data.get('Adresse (automatique)', ''),
                'router_number': row_data.get('N° de routeur'),
                'last_intervention': row_data.get('Date dernière intervention'),
                'cleaning_product': row_data.get('Produit lavant'),
            }
    
    wb.close()
    return facilities_data


def get_facility_info(excel_data: Dict[int, dict], facility_id: int) -> Optional[dict]:
    """
    Récupère les informations d'une facility depuis les données Excel
    
    Args:
        excel_data: Données parsées depuis le fichier Excel
        facility_id: ID de la facility
        
    Returns:
        dict ou None si la facility n'existe pas
    """
    return excel_data.get(facility_id)


def enrich_facility_data(facility: dict, excel_data: Dict[int, dict]) -> dict:
    """
    Enrichit les données d'une facility avec les informations du fichier Excel
    
    Args:
        facility: Données de base de la facility
        excel_data: Données parsées depuis le fichier Excel
        
    Returns:
        dict: Facility enrichie
    """
    facility_id = facility.get('facilityId')
    
    if not facility_id:
        return facility
    
    excel_info = get_facility_info(excel_data, facility_id)
    
    if excel_info:
        # Enrichir avec les données Excel
        facility['excel_data'] = excel_info
        
        # Mettre à jour certains champs si vides
        if not facility.get('facilityName') and excel_info.get('client_name'):
            facility['facilityName'] = excel_info['client_name']
            
    return facility
