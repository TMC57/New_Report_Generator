"""
Parser pour le fichier Excel "Listing clients Orsye-Wash.xlsx"
Extrait les informations nécessaires pour enrichir les rapports
"""

import openpyxl
import csv
from typing import Dict, List, Optional
from pathlib import Path


def parse_listing_clients_excel(file_path: str) -> Dict[int, dict]:
    """
    Parse le fichier Excel "Listing clients Orsye-Wash.xlsx" ou CSV
    
    Colonnes attendues:
    - Date installation
    - N° client
    - Client (automatique)
    - Groupe (automatique)
    - Date installation (duplicate?)
    - N° de zone de lavage
    - Adresse (automatique)
    - N° de routeur
    - Date dernière intervention
    - Produit lavant
    
    Returns:
        Dict[facility_id, dict]: Dictionnaire avec facility_id comme clé
    """
    
    file_path_obj = Path(file_path)
    
    # Supporter les fichiers CSV
    if file_path_obj.suffix.lower() == '.csv':
        return _parse_csv(file_path)
    
    # Fichiers Excel (.xlsx, .xls)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Lire les en-têtes
    headers = [cell.value for cell in ws[1]]
    
    # Mapping des données par facility_id (N° client)
    facilities_data = {}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):  # Skip empty rows
            continue
            
        # Créer un dictionnaire pour cette ligne
        row_data = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_data[header] = row[i]
        
        # Extraire le facility_id (N° client)
        facility_id = row_data.get('N° client')
        
        if facility_id:
            # Convertir en int si possible
            try:
                facility_id = int(facility_id)
            except (ValueError, TypeError):
                continue
            
            # Stocker les informations utiles
            facilities_data[facility_id] = {
                'client_name': row_data.get('Client (automatique)', ''),
                'group': row_data.get('Groupe (automatique)', ''),
                'installation_date': row_data.get('Date installation'),
                'zone_number': row_data.get('N° de zone de lavage'),
                'address': row_data.get('Adresse (automatique)', ''),
                'router_number': row_data.get('N° de routeur'),
                'last_intervention': row_data.get('Date dernière intervention'),
                
                # Zone 1 (principale)
                'produit_lavant': row_data.get('Produit lavant'),
                'dilution_lavant': row_data.get('Dilution lavant'),
                'couleur_buse_lavant': row_data.get('Couleur buse lavant'),
                'produit_sechant': row_data.get('Produit séchant'),
                'dilution_sechant': row_data.get('Dilution séchant'),
                'couleur_buse_sechant': row_data.get('Couleur buse séchant'),
                'autre_produit_lavant': row_data.get('Autre produit lavant'),
                'autre_dilution_lavant': row_data.get('Autre dilution lavant'),
                'autre_couleur_buse_lavant': row_data.get('Autre couleur buse lavant'),
                'produit_jantes': row_data.get('Produit jantes'),
                'dilution_jantes': row_data.get('Dilution jantes'),
                
                # Zone 2
                'produit_lavant_zone2': row_data.get('Produit lavant Zone 2'),
                'dilution_lavant_zone2': row_data.get('Dilution lavant Zone 2'),
                'couleur_buse_lavant_zone2': row_data.get('Couleur buse lavant Zone 2'),
                'produit_sechant_zone2': row_data.get('Produit séchant Zone 2'),
                'dilution_sechant_zone2': row_data.get('Dilution séchant Zone 2'),
                'couleur_buse_sechant_zone2': row_data.get('Couleur buse séchant Zone 2'),
                'autre_produit_lavant_zone2': row_data.get('Autre produit lavant Zone 2'),
                'autre_dilution_lavant_zone2': row_data.get('Autre dilution lavant Zone 2'),
                'autre_couleur_buse_lavant_zone2': row_data.get('Autre couleur buse lavant Zone 2'),
                
                # Zone 3
                'produit_lavant_zone3': row_data.get('Produit lavant Zone 3'),
                'dilution_lavant_zone3': row_data.get('Dilution lavant Zone 3'),
                'couleur_buse_lavant_zone3': row_data.get('Couleur buse lavant Zone 3'),
                'produit_sechant_zone3': row_data.get('Produit séchant Zone 3'),
                'dilution_sechant_zone3': row_data.get('Dilution séchant Zone 3'),
                'couleur_buse_sechant_zone3': row_data.get('Couleur buse séchant Zone 3'),
                
                # Zone 4
                'produit_lavant_zone4': row_data.get('Produit lavant Zone 4'),
                'dilution_lavant_zone4': row_data.get('Dilution lavant Zone 4'),
                'couleur_buse_lavant_zone4': row_data.get('Couleur buse lavant Zone 4'),
                'produit_sechant_zone4': row_data.get('Produit séchant Zone 4'),
                'dilution_sechant_zone4': row_data.get('Dilution séchant Zone 4'),
                'couleur_buse_sechant_zone4': row_data.get('Couleur buse séchant Zone 4'),
                
                # Zone 5
                'produit_lavant_zone5': row_data.get('Produit lavant Zone 5'),
                'dilution_lavant_zone5': row_data.get('Dilution lavant Zone 5'),
                'couleur_buse_lavant_zone5': row_data.get('Couleur buse lavant Zone 5'),
                'produit_sechant_zone5': row_data.get('Produit séchant Zone 5'),
                'dilution_sechant_zone5': row_data.get('Dilution séchant Zone 5'),
                'couleur_buse_sechant_zone5': row_data.get('Couleur buse séchant Zone 5'),
            }
    
    wb.close()
    return facilities_data


def _parse_csv(file_path: str) -> Dict[int, dict]:
    """
    Parse un fichier CSV avec séparateur point-virgule
    """
    facilities_data = {}
    
    # Essayer différents encodages
    encodings = ['utf-8', 'cp1252', 'latin1', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                # Détecter le séparateur
                first_line = f.readline()
                f.seek(0)
                
                delimiter = ';' if ';' in first_line else ','
                
                reader = csv.DictReader(f, delimiter=delimiter)
                
                for row in reader:
                    # Extraire le facility_id (N° client)
                    facility_id_str = row.get('N° client') or row.get('N� client')
                    
                    if not facility_id_str:
                        continue
                    
                    try:
                        facility_id = int(facility_id_str)
                    except (ValueError, TypeError):
                        continue
                    
                    # Stocker les informations utiles
                    facilities_data[facility_id] = {
                        'client_name': row.get('Client (automatique)', ''),
                        'group': row.get('Groupe (automatique)', ''),
                        'installation_date': row.get('Date installation'),
                        'zone_number': row.get('N° de zone de lavage') or row.get('N� de zone de lavage'),
                        'address': row.get('Adresse (automatique)', ''),
                        'router_number': row.get('N° de routeur') or row.get('N� de routeur'),
                        'last_intervention': row.get('Date dernière intervention') or row.get('Date derni�re intervention'),
                        
                        # Zone 1 (principale)
                        'produit_lavant': row.get('Produit lavant'),
                        'dilution_lavant': row.get('Dilution lavant'),
                        'couleur_buse_lavant': row.get('Couleur buse lavant'),
                        'produit_sechant': row.get('Produit séchant') or row.get('Produit s�chant'),
                        'dilution_sechant': row.get('Dilution séchant') or row.get('Dilution s�chant'),
                        'couleur_buse_sechant': row.get('Couleur buse séchant') or row.get('Couleur buse s�chant'),
                        'autre_produit_lavant': row.get('Autre produit lavant'),
                        'autre_dilution_lavant': row.get('Autre dilution lavant'),
                        'autre_couleur_buse_lavant': row.get('Autre couleur buse lavant'),
                        'produit_jantes': row.get('Produit jantes'),
                        'dilution_jantes': row.get('Dilution jantes'),
                        
                        # Zone 2
                        'produit_lavant_zone2': row.get('Produit lavant Zone 2'),
                        'dilution_lavant_zone2': row.get('Dilution lavant Zone 2'),
                        'couleur_buse_lavant_zone2': row.get('Couleur buse lavant Zone 2'),
                        'produit_sechant_zone2': row.get('Produit séchant Zone 2') or row.get('Produit s�chant Zone 2'),
                        'dilution_sechant_zone2': row.get('Dilution séchant Zone 2') or row.get('Dilution s�chant Zone 2'),
                        'couleur_buse_sechant_zone2': row.get('Couleur buse séchant Zone 2') or row.get('Couleur buse s�chant Zone 2'),
                        'autre_produit_lavant_zone2': row.get('Autre produit lavant Zone 2'),
                        'autre_dilution_lavant_zone2': row.get('Autre dilution lavant Zone 2'),
                        'autre_couleur_buse_lavant_zone2': row.get('Autre couleur buse lavant Zone 2'),
                        
                        # Zone 3
                        'produit_lavant_zone3': row.get('Produit lavant Zone 3'),
                        'dilution_lavant_zone3': row.get('Dilution lavant Zone 3'),
                        'couleur_buse_lavant_zone3': row.get('Couleur buse lavant Zone 3'),
                        'produit_sechant_zone3': row.get('Produit séchant Zone 3') or row.get('Produit s�chant Zone 3'),
                        'dilution_sechant_zone3': row.get('Dilution séchant Zone 3') or row.get('Dilution s�chant Zone 3'),
                        'couleur_buse_sechant_zone3': row.get('Couleur buse séchant Zone 3') or row.get('Couleur buse s�chant Zone 3'),
                        
                        # Zone 4
                        'produit_lavant_zone4': row.get('Produit lavant Zone 4'),
                        'dilution_lavant_zone4': row.get('Dilution lavant Zone 4'),
                        'couleur_buse_lavant_zone4': row.get('Couleur buse lavant Zone 4'),
                        'produit_sechant_zone4': row.get('Produit séchant Zone 4') or row.get('Produit s�chant Zone 4'),
                        'dilution_sechant_zone4': row.get('Dilution séchant Zone 4') or row.get('Dilution s�chant Zone 4'),
                        'couleur_buse_sechant_zone4': row.get('Couleur buse séchant Zone 4') or row.get('Couleur buse s�chant Zone 4'),
                        
                        # Zone 5
                        'produit_lavant_zone5': row.get('Produit lavant Zone 5'),
                        'dilution_lavant_zone5': row.get('Dilution lavant Zone 5'),
                        'couleur_buse_lavant_zone5': row.get('Couleur buse lavant Zone 5'),
                        'produit_sechant_zone5': row.get('Produit séchant Zone 5') or row.get('Produit s�chant Zone 5'),
                        'dilution_sechant_zone5': row.get('Dilution séchant Zone 5') or row.get('Dilution s�chant Zone 5'),
                        'couleur_buse_sechant_zone5': row.get('Couleur buse séchant Zone 5') or row.get('Couleur buse s�chant Zone 5'),
                    }
                
                return facilities_data
                
        except UnicodeDecodeError:
            continue
    
    # Si aucun encodage n'a fonctionné
    raise ValueError(f"Impossible de lire le fichier CSV avec les encodages: {encodings}")


def get_facility_info(excel_data: Dict[int, dict], facility_id: int) -> Optional[dict]:
    """
    Récupère les informations d'une facility depuis les données Excel
    
    Args:
        excel_data: Données parsées depuis le fichier Excel
        facility_id: ID de la facility
        
    Returns:
        dict ou None si la facility n'existe pas
    """
    return excel_data.get(facility_id)


def enrich_facility_data(facility: dict, excel_data: Dict[int, dict]) -> dict:
    """
    Enrichit les données d'une facility avec les informations du fichier Excel
    
    Args:
        facility: Données de base de la facility
        excel_data: Données parsées depuis le fichier Excel
        
    Returns:
        dict: Facility enrichie
    """
    facility_id = facility.get('facilityId')
    
    if not facility_id:
        return facility
    
    excel_info = get_facility_info(excel_data, facility_id)
    
    if excel_info:
        # Enrichir avec les données Excel
        facility['excel_data'] = excel_info
        
        # Mettre à jour certains champs si vides
        if not facility.get('facilityName') and excel_info.get('client_name'):
            facility['facilityName'] = excel_info['client_name']
            
    return facility
